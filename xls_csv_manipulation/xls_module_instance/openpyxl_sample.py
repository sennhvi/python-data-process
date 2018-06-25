from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])
ws.append([100])
ws.append([200])
# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
